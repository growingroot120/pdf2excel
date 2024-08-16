import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from PIL import Image, ImageStat
import re
import os
import glob


# Define a function to convert RGB to Hex
def rgb_to_hex(color):
    red = (color >> 16) & 0xFF
    green = (color >> 8) & 0xFF
    blue = color & 0xFF
    hex_color = f"{red:02X}{green:02X}{blue:02X}"
    return hex_color


# Write text to Excel with formatting
def read_pdf_file_page(blocks):
    row_number = 1
    lines = []
    for block in blocks:
        block_array = []
        col_number = 1
        if "lines" in block:  # Ensure the block contains lines of text
            for line in block["lines"]:
                for span in line["spans"]:
                    text = span["text"]
                    color = span["color"]  # Color in 8-bit RGBA format
                    font_size = span["size"]

                    # Calculate the column width
                    column_width = max(len(text), 10)

                    element_array = [text, color, font_size, column_width]
                    block_array.append(element_array)

                    # # Add text to Excel cell
                    # cell = sheet.cell(row=row_number, column=col_number, value=text)

                    # # Set font size and color
                    # cell.font = Font(size=font_size, color=rgb_to_hex(color))

                    # # Adjust column width
                    # sheet.column_dimensions[get_column_letter(1)].width = column_width
                    col_number += 1
        lines.append(block_array)
        row_number += 1

    return lines


def is_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def separate_lines(lines, color_set):
    separate_lines = lines
    is_color_set = color_set
    for index in range(2, len(lines)):
        if index == 2:
            col_BM = lines[index][-1]
            separate_lines[index].remove(col_BM)
            separate_lines[index].insert(0, col_BM)
            col_settle = separate_lines[index][-2]
            separate_lines[index].remove(col_settle)
            separate_lines[index].insert(1, col_settle)
            space_col = separate_lines[index][2].copy()
            space_col[0] = "#"
            separate_lines[index].insert(3, space_col)
            separate_lines[index].insert(5, space_col)
            separate_lines[index].insert(7, space_col)
            separate_lines[index].insert(9, space_col)
            separate_lines[index].insert(11, space_col)
            separate_lines[index].insert(13, space_col)
            separate_lines[index].insert(15, space_col)
            separate_lines[index].insert(17, space_col)
        else:

            if len(separate_lines[index]) > 4:
                # Set first column value
                column_zero = separate_lines[index][1]
                column_first = separate_lines[index][2]
                separate_lines[index].remove(column_zero)
                separate_lines[index].remove(column_first)
                separate_lines[index].insert(0, column_zero)
                separate_lines[index].insert(1, column_first)

                # Set second column value
                col_settle = separate_lines[index][-1]
                separate_lines[index].pop()
                separate_lines[index].insert(3, col_settle)

                # Separate number and bracket
                n = 0
                for brac_index in range(4, len(separate_lines[index])):
                    # if (brac_index + n) > (len(separate_lines[index]) - 1):
                    #     break

                    complex_value = separate_lines[index][brac_index + n][0]
                    if (
                        "[" in complex_value
                        and "]" in complex_value
                        and "." in complex_value
                        and len(complex_value) > 6
                    ):

                        main_value, bracketed_value = re.split(
                            r"\s*\[\s*", complex_value
                        )
                        bracketed_value = bracketed_value.rstrip("]")
                        bracketed_value = bracketed_value
                        complex_element = separate_lines[index][brac_index + n]
                        separate_lines[index].remove(complex_element)
                        complex_element[0] = main_value
                        main_element = complex_element.copy()
                        complex_element[0] = bracketed_value
                        bracket_element = complex_element
                        separate_lines[index].insert(brac_index + n, main_element)
                        separate_lines[index].insert(
                            brac_index + n + 1, bracket_element
                        )
                        n += 1

                    else:
                        if brac_index + n + 1 > len(separate_lines[index]) - 1:
                            break
                        if "[" in separate_lines[index][brac_index + n + 1][0]:
                            # n += 1
                            continue
                        elif "[" in separate_lines[index][brac_index + n][0]:
                            continue
                        else:

                            if brac_index + n > len(separate_lines[index]) - 1:
                                break
                            elif ":" in separate_lines[index][brac_index + n][0]:
                                continue
                            elif "[" not in separate_lines[index][brac_index + n][0]:
                                continue
                            else:
                                separate_lines[index].insert(brac_index + n, space_col)
                                n += 1

                # Set value in 1600m to 200m columns
                col_Num_space = separate_lines[index][4].copy()
                col_Num_space[0] = "#"
                if is_color_set[1] == 1:
                    col_1400m_bracket = separate_lines[index][-1]
                    separate_lines[index].pop()
                    col_1400m_Num = separate_lines[index][-1]
                    separate_lines[index].pop()
                    separate_lines[index].insert(4, col_1400m_Num)
                    separate_lines[index].insert(5, col_1400m_bracket)
                else:
                    separate_lines[index].insert(4, col_Num_space)
                    separate_lines[index].insert(5, col_Num_space)

                if is_color_set[0] == 1:
                    col_1600m_bracket = separate_lines[index][-1]
                    separate_lines[index].pop()
                    col_1600m_Num = separate_lines[index][-1]
                    separate_lines[index].pop()
                    separate_lines[index].insert(4, col_1600m_Num)
                    separate_lines[index].insert(5, col_1600m_bracket)
                else:
                    separate_lines[index].insert(4, col_Num_space)
                    separate_lines[index].insert(5, col_Num_space)

                if is_color_set[4] == 1:
                    col_800m_bracket = separate_lines[index][-1]
                    separate_lines[index].pop()
                    col_800m_Num = separate_lines[index][-1]
                    separate_lines[index].pop()
                    separate_lines[index].insert(8, col_800m_Num)
                    separate_lines[index].insert(9, col_800m_bracket)
                else:
                    separate_lines[index].insert(8, col_Num_space)
                    separate_lines[index].insert(9, col_Num_space)

                if is_color_set[3] == 1:
                    col_1000m_bracket = separate_lines[index][-1]
                    separate_lines[index].pop()
                    col_1000m_Num = separate_lines[index][-1]
                    separate_lines[index].pop()
                    separate_lines[index].insert(8, col_1000m_Num)
                    separate_lines[index].insert(9, col_1000m_bracket)
                else:
                    separate_lines[index].insert(8, col_Num_space)
                    separate_lines[index].insert(9, col_Num_space)

                if is_color_set[2] == 1:
                    col_1200m_bracket = separate_lines[index][-1]
                    separate_lines[index].pop()
                    col_1200m_Num = separate_lines[index][-1]
                    separate_lines[index].pop()
                    separate_lines[index].insert(8, col_1200m_Num)
                    separate_lines[index].insert(9, col_1200m_bracket)
                else:
                    separate_lines[index].insert(8, col_Num_space)
                    separate_lines[index].insert(9, col_Num_space)

                if is_color_set[5] == 0:
                    separate_lines[index].insert(12, col_Num_space)
                    separate_lines[index].insert(12, col_Num_space)

                if is_color_set[6] == 0:
                    separate_lines[index].insert(14, col_Num_space)
                    separate_lines[index].insert(14, col_Num_space)

                if is_color_set[7] == 0:
                    separate_lines[index].insert(16, col_Num_space)
                    separate_lines[index].insert(16, col_Num_space)

                # Set finish to fixed
                if "#" in separate_lines[index][-1][0]:
                    separate_lines[index].pop()
                finish_line = separate_lines[index][-1].copy()
                finish_text = finish_line[0]

                if ":" in finish_text:
                    minutes, seconds = finish_text.split(":")
                    if len(str(seconds)) == 4:
                        seconds = float(seconds)
                    seconds = f"{float(seconds):05.2f}"
                    finish_text = f"{int(minutes):02}:{seconds}"
                else:
                    if is_float(finish_text):
                        finish_text = float(finish_text)
                        minutes = int(finish_text // 60)
                        seconds = finish_text % 60
                        finish_text = f"{minutes:02}:{seconds:05.2f}"
                    else:
                        finish_text

                separate_lines[index][-1][0] = finish_text
                if (
                    separate_lines[index][-2][0] == "#"
                    and ":" in separate_lines[index][-1][0]
                ):
                    separate_lines[index].pop(-2)


    return separate_lines


def make_table_column_names(pre_lines):

    # Add column names
    date_text = pre_lines[0][0].copy()
    table_column_names = []
    date_text[0] = "DATE"
    table_column_names.append(date_text)
    track_text = date_text.copy()
    track_text[0] = "TRACK"
    table_column_names.append(track_text)
    race_number_text = date_text.copy()
    race_number_text[0] = "Race number"
    table_column_names.append(race_number_text)
    race_distance_text = date_text.copy()
    race_distance_text[0] = "RACE DISTANCE"
    table_column_names.append(race_distance_text)
    race_name_text = date_text.copy()
    race_name_text[0] = "RACE NAME"
    table_column_names.append(race_name_text)
    class_text = date_text.copy()
    class_text[0] = "CLASS"
    table_column_names.append(class_text)
    bias_text = date_text.copy()
    bias_text[0] = "BIAS"
    table_column_names.append(bias_text)
    finsh_pos_text = date_text.copy()
    finsh_pos_text[0] = "finish pos"
    table_column_names.append(finsh_pos_text)
    tab_number_text = date_text.copy()
    tab_number_text[0] = "TAB NUMBER"
    table_column_names.append(tab_number_text)
    horse_name_text = date_text.copy()
    horse_name_text[0] = "horse name"
    table_column_names.append(horse_name_text)
    suffix_text = date_text.copy()
    suffix_text[0] = "suffix"
    table_column_names.append(suffix_text)

    for index in range(2, 19, 2):
        x = pre_lines[2][index].copy()
        table_column_names.append(x)

    settle_pir_text = date_text.copy()
    settle_pir_text[0] = "settle PIR"
    table_column_names.append(settle_pir_text)

    for index in range(8):
        pir_number = 1600 - 200 * index
        level_pir_text = date_text.copy()
        level_pir_text[0] = f"{pir_number} pir"
        table_column_names.append(level_pir_text)

    table_lines.append(table_column_names)


def make_table_contents(process_lines):
    # Set date and track column values
    process_date_track_line = process_lines[0][0].copy()
    process_date_track_text = process_date_track_line[0]
    date_track_text = process_date_track_text.split(" at ")
    date_text_array = date_track_text[0].split()
    date_text = date_text_array[1] + "-" + date_text_array[2] + "-" + date_text_array[3]
    process_date_track_line[0] = date_text
    process_track_line = process_date_track_line.copy()
    process_track_line[0] = date_track_text[1].strip()

    # Set race number, race distance, and race name column values
    race_process_line = process_lines[1][0].copy()
    race_number_distance_text = race_process_line[0]
    race_num_dis_array = race_number_distance_text.split()
    race_num_text = race_num_dis_array[1]
    race_num_text = race_num_text.split(",")[0]
    race_num_line = race_process_line.copy()
    race_num_line[0] = race_num_text
    race_distance_text = race_num_dis_array[2]
    race_distance_text = race_distance_text.split("m")[0]
    race_distance_line = race_process_line.copy()
    race_distance_line[0] = race_distance_text

    for index, line in enumerate(process_lines, start=3):
        table_row_line = []
        if len(process_lines[index]) > 4:
            # Add track column value
            table_row_line.append(process_date_track_line)
            table_row_line.append(process_track_line)
            # Add race number column value
            table_row_line.append(race_num_line)
            table_row_line.append(race_distance_line)
            # Add race name column value
            race_name_element = process_lines[1][1].copy()
            table_row_line.append(race_name_element)
            # Add class column value
            class_element = process_lines[2][0].copy()
            table_row_line.append(class_element)
            # Add bias column value
            bias_element = process_lines[-3][-1].copy()
            table_row_line.append(bias_element)
            # Add finish pos column value
            finish_pos_element = process_lines[index][0].copy()
            table_row_line.append(finish_pos_element)
            # Add tab number column value
            tab_number_element = process_lines[index][1].copy()
            table_row_line.append(tab_number_element)

            # Add horse name and suffix column values
            horse_name_element = process_lines[index][2].copy()
            horse_name_text = horse_name_element[0]
            match = re.match(r"(\w+)\s\((\w+)\)", horse_name_text)

            if match:
                part1 = match.group(1)
                part2 = match.group(2)
                horse_1_process_element = horse_name_element.copy()
                horse_1_process_element[0] = part1
                horse_2_process_element = horse_name_element.copy()
                horse_2_process_element[0] = part2
                table_row_line.append(horse_1_process_element)
                table_row_line.append(horse_2_process_element)
            else:
                horse_process_element = process_lines[index][2].copy()
                table_row_line.append(horse_name_element)
                horse_process_element[0] = ""
                table_row_line.append(horse_process_element)

            # Add 1600m ~ 200m columns values
            for time_index in range(4, len(process_lines[index]) - 1, 2):
                time_element = process_lines[index][time_index].copy()
                table_row_line.append(time_element)

            # Add finish column value
            finish_time_element = process_lines[index][-1].copy()
            table_row_line.append(finish_time_element)

            # Add settle PIR column value
            settle_pir_element = process_lines[index][3].copy()
            table_row_line.append(settle_pir_element)

            for pir_index in range(5, len(process_lines[index]) - 1, 2):
                pir_element = process_lines[index][pir_index].copy()
                pir_element_text = pir_element[0]
                if "[" in pir_element_text:
                    bracketed_value = re.split(r"\s*\[\s*", pir_element_text)[1]
                    bracketed_value = bracketed_value.rstrip("]")
                    pir_element[0] = bracketed_value
                table_row_line.append(pir_element)

            if len(table_row_line) < 29:
                i = len(table_row_line)
                correct_col = table_row_line[2].copy()
                correct_col[0] = "#"
                while i < 29:
                    table_row_line.append(correct_col)
                    i += 1

            table_lines.append(table_row_line)
        else:
            print("-----------------------Processing...")
            break


def write_to_excel(table_lines, sheet):
    sheet.freeze_panes = "A2"
    for row_number, row in enumerate(table_lines, start=1):
        for col_number, element in enumerate(row, start=1):
            text, color, font_size, column_width = element
            cell = sheet.cell(row=row_number, column=(col_number), value=text)
            sheet.column_dimensions[get_column_letter(col_number)].width = column_width
            cell.font = Font(size=font_size, color=rgb_to_hex(color))
            white_color = 0xFFFFFF
            white_int = rgb_to_hex(white_color)
            white_fill = PatternFill(
                start_color=white_int,
                end_color=white_int,
                fill_type="solid",
            )
            if row_number == 1:
                if col_number > 11 and col_number < 21:
                    pink_color = 0xFFC0CB
                    pink_int = rgb_to_hex(pink_color)
                    pink_fill = PatternFill(
                        start_color=pink_int, end_color=pink_int, fill_type="solid"
                    )
                    cell.fill = pink_fill
                elif col_number > 20:
                    green_color = 0x32CD32
                    green_int = rgb_to_hex(green_color)
                    green_fill = PatternFill(
                        start_color=green_int,
                        end_color=green_int,
                        fill_type="solid",
                    )
                    cell.fill = green_fill
                else:

                    cell.fill = white_fill
            else:
                # Add text to Excel cell

                cell.fill = white_fill


if __name__ == "__main__":

    pdf_folder = os.path.dirname(os.path.abspath(__file__))
    pdf_files = glob.glob(os.path.join(pdf_folder, "*.pdf"))
    start_row_num = 1
    pdf_number = 0
    for pdf_file in pdf_files:
        print(f"Processing {pdf_file}...")

        # Load the PDF file
        doc = fitz.open(pdf_file)
        table_lines = []
        # Create an Excel workbook and sheet
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "PDF Content"

        for page_number in range(len(doc)):
            print(f"------Processing {page_number} page...")
            page = doc.load_page(page_number)
            blocks = page.get_text("dict")["blocks"]

            rect = fitz.Rect(275, 80, 818, 95)
            # rect = fitz.Rect(275, 65, 818, 80)
            # # Extract the image
            pix = page.get_pixmap(clip=rect)

            # Save the full image
            # output_image_path = "extracted_image.png"
            # pix.save(output_image_path)

            image = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            # Divide the image into 5 columns
            num_columns = 9
            image_width, image_height = image.size
            column_width = image_width // num_columns

            threshold = 249  # Adjust this value as needed
            color_set = []
            for i in range(num_columns):
                left = i * column_width
                right = (i + 1) * column_width if i < num_columns - 1 else image_width
                column_image = image.crop((left, 0, right, image_height))

                # Save the extracted column image as a PNG file
                # column_image_path = f"{os.path.splitext(pdf_file)[0]}_page_{page_number + 1}_column_{i + 1}.png"
                # column_image.save(column_image_path)

                stat = ImageStat.Stat(column_image)
                avg_color = stat.mean[:3]
                avg_color_int = (
                    int(avg_color[0]),
                    int(avg_color[1]),
                    int(avg_color[2]),
                )

                # Determine if the average color is white
                is_white = all(c > threshold for c in avg_color_int)
                if is_white:
                    color_set.append(0)
                else:
                    color_set.append(1)

            lines = read_pdf_file_page(blocks)
            separated_lines = separate_lines(lines, color_set)
            if page_number == 0:
                make_table_column_names(separated_lines)
            make_table_contents(separated_lines)

            start_row_num += len(separated_lines) + 1
        write_to_excel(table_lines, sheet)
        output_excel_path = f"{os.path.splitext(pdf_file)[0]}_extracted.xlsx"
        workbook.save(output_excel_path)

    print(f"Text from the first page has been written to {output_excel_path}")
